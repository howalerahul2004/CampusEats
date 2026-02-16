from flask import Flask, render_template as _flask_render_template, request, redirect, url_for, session, jsonify, flash, send_file
import mysql.connector
from mysql.connector import Error
from datetime import datetime, timedelta, date, time as dtime
import os
import json
import random
from functools import wraps
from io import BytesIO
import base64
import uuid
from werkzeug.security import generate_password_hash, check_password_hash
import re

app = Flask(__name__)
app.secret_key = 'your_secret_key_here_change_in_production'


# Render template wrapper: normalize Decimal/timedelta for templates
def _normalize_value(v):
    """Recursively convert DB-returned types to template-friendly types.

    - decimal.Decimal -> float
    - datetime.timedelta -> datetime.time
    - list/tuple/dict -> recurse
    """
    from decimal import Decimal
    from datetime import timedelta, datetime

    if v is None:
        return v

    # Decimal -> float
    try:
        if isinstance(v, Decimal):
            return float(v)
    except Exception:
        pass

    # timedelta -> time
    try:
        if isinstance(v, timedelta):
            return (datetime.min + v).time()
    except Exception:
        pass

    # dict -> recurse
    if isinstance(v, dict):
        return {k: _normalize_value(val) for k, val in v.items()}

    # list/tuple -> recurse
    if isinstance(v, list):
        return [_normalize_value(x) for x in v]
    if isinstance(v, tuple):
        return tuple(_normalize_value(x) for x in v)

    return v


def render_template(template_name, **context):
    """Wrapper around Flask's render_template that normalizes context values.

    Use this in place of Flask's render_template across the app.
    """
    normalized = {k: _normalize_value(v) for k, v in context.items()}
    return _flask_render_template(template_name, **normalized)



# --- Jinja2 safe formatting filters -------------------------------------------------
def format_datetime(value, fmt='%b %d, %Y %I:%M %p'):
    """Safely format datetime/date/time/timedelta/ISO-string values for templates.

    - datetime, date, time: use strftime
    - timedelta: add to datetime.min and format
    - ISO strings: try to parse with fromisoformat(), otherwise return as-is
    - None: return empty string
    """
    from datetime import datetime, date, time, timedelta

    if value is None:
        return ''

    # datetime.date/time/datetime
    try:
        if isinstance(value, datetime) or isinstance(value, date) or isinstance(value, time):
            return value.strftime(fmt)

        if isinstance(value, timedelta):
            return (datetime.min + value).strftime(fmt)

        if isinstance(value, str):
            # try ISO parse
            try:
                # try datetime
                dt = datetime.fromisoformat(value)
                return dt.strftime(fmt)
            except Exception:
                # return raw string
                return value

        # Fallback
        return str(value)
    except Exception:
        return str(value)


app.jinja_env.filters['format_datetime'] = format_datetime

# ----------------------------------------------------------------------------------


@app.context_processor
def inject_user():
    return dict(current_user=({'user_id': session.get('user_id'), 'user_name': session.get('user_name'), 'user_type': session.get('user_type')} if session.get('user_id') else None), subscription_status=session.get('subscription_status', 'inactive'))


# Database Configuration
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': '',  # Change this
    'database': 'campus_eats'
}

# Database Connection Helper
def get_db_connection():
    try:
        connection = mysql.connector.connect(**DB_CONFIG)
        return connection
    except Error as e:
        print(f"Error connecting to MySQL: {e}")
        return None

# Login Required Decorator
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Admin Login Required Decorator
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin_id' not in session:
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function

# Helper function to create notification
def create_notification(user_id, order_id, title, message, notif_type='order'):
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO notifications (user_id, order_id, title, message, type)
            VALUES (%s, %s, %s, %s, %s)
        """, (user_id, order_id, title, message, notif_type))
        conn.commit()
        cursor.close()
        conn.close()

# Helper function to update special menu daily - ensures only one special per 24 hours from meals/snacks
def update_daily_special():
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor(dictionary=True)
        
        # Check if there's already a special for today
        cursor.execute("""
            SELECT food_id, category_id 
            FROM food_items 
            WHERE special_date = CURDATE() AND is_special = TRUE LIMIT 1
        """)
        today = cursor.fetchone()

        # Reset all is_special flags first
        cursor.execute("UPDATE food_items SET is_special = FALSE")
        
        if today and today.get('food_id'):
            # If there's already a special for today, just keep it marked
            cursor.execute("UPDATE food_items SET is_special = TRUE WHERE food_id = %s", (today['food_id'],))
        else:
            # Pick one random available food from categories 1 (Snacks) or 2 (Meals) only
            cursor.execute("""
                SELECT food_id 
                FROM food_items
                WHERE is_available = TRUE 
                  AND (special_date IS NULL OR special_date != CURDATE())
                  AND category_id IN (1, 2)  -- Snacks (1) or Meals (2) only
                ORDER BY RAND() LIMIT 1
            """)
            result = cursor.fetchone()
            if result:
                # Set new special for today
                cursor.execute("""
                    UPDATE food_items
                    SET is_special = TRUE, special_date = CURDATE()
                    WHERE food_id = %s
                """, (result['food_id'],))
        
        conn.commit()
        cursor.close()
        conn.close()

# ==================== USER ROUTES ====================

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/about')
def about():
    return render_template('about.html')

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        unique_user_id = request.form['unique_user_id']
        name = request.form['name']
        email = request.form['email']
        password = request.form['password']
        confirm_password = request.form.get('confirm_password')
        phone = request.form['phone']
        user_type = request.form['user_type']

        # Server-side validations
        # 1) password and confirm match
        if not confirm_password or password != confirm_password:
            flash('Passwords do not match.', 'error')
            return render_template('signup.html')

        # 2) unique_user_id: first char letter, max 12 chars, rest alnum
        if not re.match(r'^[A-Za-z][A-Za-z0-9]{0,11}$', unique_user_id):
            flash('Unique User ID must start with a letter and be up to 12 alphanumeric characters.', 'error')
            return render_template('signup.html')

        # 3) phone must be digits only and max 10 digits
        phone_digits = ''.join(ch for ch in phone if ch.isdigit())
        if not phone_digits or len(phone_digits) > 10:
            flash('Phone number must contain up to 10 digits only.', 'error')
            return render_template('signup.html')
        # normalize phone to digits-only (store without formatting)
        phone = phone_digits
        
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            try:
                # Hash the password before storing
                hashed_password = generate_password_hash(password)
                cursor.execute("""
                    INSERT INTO users (unique_user_id, name, email, password, phone, user_type)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (unique_user_id, name, email, hashed_password, phone, user_type))
                user_id = cursor.lastrowid
                
                # Create default preferences
                cursor.execute("""
                    INSERT INTO user_preferences (user_id, theme) VALUES (%s, 'light')
                """, (user_id,))
                
                conn.commit()
                flash('Registration successful! Please login.', 'success')
                return redirect(url_for('login'))
            except Error as e:
                flash(f'Registration failed: {str(e)}', 'error')
            finally:
                cursor.close()
                conn.close()
    
    return render_template('signup.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("""
                SELECT u.*, up.theme 
                FROM users u
                LEFT JOIN user_preferences up ON u.user_id = up.user_id
                WHERE u.email = %s
            """, (email,))
            user = cursor.fetchone()
            cursor.close()
            conn.close()
            
            if user and check_password_hash(user.get('password', ''), password):
                session['user_id'] = user['user_id']
                session['user_name'] = user['name']
                session['user_type'] = user['user_type']
                session['theme'] = user['theme'] if user['theme'] else 'light'
                # Keep subscription status in session so templates can show discounts/offers
                session['subscription_status'] = user.get('subscription_status', 'inactive')
                return redirect(url_for('menu'))
            else:
                flash('Invalid credentials!', 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))


def _ensure_password_resets_table():
    """Create password_resets table if it doesn't exist (safe to call on startup)."""
    conn = get_db_connection()
    if not conn:
        return
    cursor = conn.cursor()
    try:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS password_resets (
                id INT AUTO_INCREMENT PRIMARY KEY,
                user_id INT NOT NULL,
                token VARCHAR(128) NOT NULL,
                expires_at DATETIME NOT NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX (token),
                INDEX (user_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """)
        conn.commit()
    except Exception:
        # non-fatal in case of permission issues; routes will handle errors
        pass
    finally:
        cursor.close()
        conn.close()


@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = request.form.get('email')
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT user_id, name FROM users WHERE email = %s", (email,))
            user = cursor.fetchone()
            if not user:
                cursor.close()
                conn.close()
                flash('If that email exists in our system, a reset link has been sent.', 'info')
                return redirect(url_for('forgot_password'))

            token = uuid.uuid4().hex
            expires_at = datetime.now() + timedelta(hours=1)

            # Ensure table exists and insert token
            _ensure_password_resets_table()
            try:
                cursor.execute("INSERT INTO password_resets (user_id, token, expires_at) VALUES (%s, %s, %s)", (user['user_id'], token, expires_at))
                conn.commit()
            except Exception:
                # fall back silently
                pass
            finally:
                cursor.close()
                conn.close()

            # In production you'd email this link. For now show link via flash (developer mode).
            reset_link = url_for('reset_password', token=token, _external=True)
            flash('Password reset link (for development): ' + reset_link, 'info')
            return redirect(url_for('login'))

    return render_template('forgot_password.html')


@app.route('/reset_password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    conn = get_db_connection()
    if not conn:
        flash('Server error. Try again later.', 'error')
        return redirect(url_for('login'))

    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM password_resets WHERE token = %s", (token,))
    row = cursor.fetchone()

    if not row or row['expires_at'] < datetime.now():
        # invalid or expired
        try:
            if row:
                # cleanup
                cursor.execute("DELETE FROM password_resets WHERE token = %s", (token,))
                conn.commit()
        except Exception:
            pass
        cursor.close()
        conn.close()
        flash('This reset link is invalid or has expired.', 'error')
        return redirect(url_for('forgot_password'))

    if request.method == 'POST':
        new_password = request.form.get('password')
        if not new_password or len(new_password) < 4:
            flash('Please provide a stronger password (min 4 chars).', 'error')
            return render_template('reset_password.html', token=token)

        try:
            # Hash the new password before storing
            hashed = generate_password_hash(new_password)
            cursor.execute("UPDATE users SET password = %s WHERE user_id = %s", (hashed, row['user_id']))
            cursor.execute("DELETE FROM password_resets WHERE token = %s", (token,))
            conn.commit()
        except Exception as e:
            flash('Failed to reset password: ' + str(e), 'error')
            cursor.close()
            conn.close()
            return render_template('reset_password.html', token=token)

        cursor.close()
        conn.close()
        flash('Password has been reset. Please login with your new password.', 'success')
        return redirect(url_for('login'))

    cursor.close()
    conn.close()
    return render_template('reset_password.html', token=token)


@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    user_id = session['user_id']
    conn = get_db_connection()
    if not conn:
        flash('Server error', 'error')
        return redirect(url_for('menu'))

    cursor = conn.cursor(dictionary=True)
    if request.method == 'POST':
        name = request.form.get('name')
        phone = request.form.get('phone')
        password = request.form.get('password')

        try:
            if password:
                # hash password before updating
                hashed_pw = generate_password_hash(password)
                cursor.execute("UPDATE users SET name = %s, phone = %s, password = %s WHERE user_id = %s", (name, phone, hashed_pw, user_id))
            else:
                cursor.execute("UPDATE users SET name = %s, phone = %s WHERE user_id = %s", (name, phone, user_id))
            conn.commit()
            session['user_name'] = name
            flash('Profile updated successfully.', 'success')
        except Exception as e:
            flash('Failed to update profile: ' + str(e), 'error')
        finally:
            cursor.close()
            conn.close()
        return redirect(url_for('profile'))

    # GET
    cursor.execute("SELECT user_id, name, email, phone FROM users WHERE user_id = %s", (user_id,))
    user = cursor.fetchone()
    cursor.close()
    conn.close()
    return render_template('profile.html', user=user)

@app.route('/menu')
@login_required
def menu():
    # Update daily special
    update_daily_special()
    
    conn = get_db_connection()
    categories = []
    foods = []
    favorites = []
    
    if conn:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM categories")
        categories = cursor.fetchall()
        
        cursor.execute("""
            SELECT f.*, c.category_name,
                   (SELECT COUNT(*) FROM favorites WHERE food_id = f.food_id AND user_id = %s) as is_favorite
            FROM food_items f 
            LEFT JOIN categories c ON f.category_id = c.category_id 
            WHERE f.is_available = TRUE
            ORDER BY f.is_special DESC, c.category_name, f.name
        """, (session['user_id'],))
        foods = cursor.fetchall()
        
        cursor.execute("""
            SELECT food_id FROM favorites WHERE user_id = %s
        """, (session['user_id'],))
        favorites = [f['food_id'] for f in cursor.fetchall()]
        
        cursor.close()
        conn.close()
    
    return render_template('menu.html', categories=categories, foods=foods, favorites=favorites)

@app.route('/add_to_cart', methods=['POST'])
@login_required
def add_to_cart():
    food_id = request.form['food_id']
    user_id = session['user_id']
    
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        try:
            cursor.execute("""
                INSERT INTO cart (user_id, food_id, quantity) 
                VALUES (%s, %s, 1)
                ON DUPLICATE KEY UPDATE quantity = quantity + 1
            """, (user_id, food_id))
            conn.commit()
            return jsonify({'success': True, 'message': 'Added to cart!'})
        except Error as e:
            return jsonify({'success': False, 'message': str(e)})
        finally:
            cursor.close()
            conn.close()
    
    return jsonify({'success': False, 'message': 'Database connection failed'})

@app.route('/toggle_favorite', methods=['POST'])
@login_required
def toggle_favorite():
    food_id = request.form['food_id']
    user_id = session['user_id']
    
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor(dictionary=True)
        try:
            # Check if already favorite
            cursor.execute("""
                SELECT * FROM favorites WHERE user_id = %s AND food_id = %s
            """, (user_id, food_id))
            existing = cursor.fetchone()
            
            if existing:
                cursor.execute("""
                    DELETE FROM favorites WHERE user_id = %s AND food_id = %s
                """, (user_id, food_id))
                message = 'Removed from favorites'
                is_favorite = False
            else:
                cursor.execute("""
                    INSERT INTO favorites (user_id, food_id) VALUES (%s, %s)
                """, (user_id, food_id))
                message = 'Added to favorites'
                is_favorite = True
            
            conn.commit()
            return jsonify({'success': True, 'message': message, 'is_favorite': is_favorite})
        except Error as e:
            return jsonify({'success': False, 'message': str(e)})
        finally:
            cursor.close()
            conn.close()
    
    return jsonify({'success': False, 'message': 'Database connection failed'})

@app.route('/favorites')
@login_required
def favorites():
    user_id = session['user_id']
    conn = get_db_connection()
    favorite_foods = []
    
    if conn:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT f.*, c.category_name
            FROM favorites fav
            JOIN food_items f ON fav.food_id = f.food_id
            LEFT JOIN categories c ON f.category_id = c.category_id
            WHERE fav.user_id = %s AND f.is_available = TRUE
        """, (user_id,))
        favorite_foods = cursor.fetchall()
        cursor.close()
        conn.close()
    
    return render_template('favorites.html', foods=favorite_foods)

@app.route('/order_favorites', methods=['POST'])
@login_required
def order_favorites():
    user_id = session['user_id']
    
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor(dictionary=True)
        
        # Get all favorite items
        cursor.execute("""
            SELECT f.food_id, f.price
            FROM favorites fav
            JOIN food_items f ON fav.food_id = f.food_id
            WHERE fav.user_id = %s AND f.is_available = TRUE
        """, (user_id,))
        favorites = cursor.fetchall()
        
        if not favorites:
            flash('No favorites to order!', 'error')
            return redirect(url_for('favorites'))
        
        # Calculate total
        total_amount = sum(item['price'] for item in favorites)
        # ensure numeric type is compatible with template math (convert Decimal -> float)
        try:
            from decimal import Decimal
            if isinstance(total_amount, Decimal):
                total_amount = float(total_amount)
        except Exception:
            pass
        
        # Get user type for priority
        cursor.execute("SELECT user_type FROM users WHERE user_id = %s", (user_id,))
        user = cursor.fetchone()
        priority = 'high' if user and user.get('user_type') == 'staff' else 'normal'

        # Create order
        cursor.execute("""
            INSERT INTO orders (user_id, total_amount, payment_method, payment_status, status, priority)
            VALUES (%s, %s, 'Quick Order', 'Paid', 'placed', %s)
        """, (user_id, total_amount, priority))
        order_id = cursor.lastrowid

        # Add order items
        for item in favorites:
            cursor.execute("""
                INSERT INTO order_items (order_id, food_id, quantity, price)
                VALUES (%s, %s, 1, %s)
            """, (order_id, item['food_id'], item['price']))

        conn.commit()

        # Create notification
        create_notification(user_id, order_id, 'Order Placed', 
                          f'Your favorite items order #{order_id} has been placed successfully!')

        cursor.close()
        conn.close()

        flash('Favorites ordered successfully!', 'success')
        return redirect(url_for('orders'))

    flash('Order failed!', 'error')
    return redirect(url_for('favorites'))

@app.route('/cart')
@login_required
def cart():
    user_id = session['user_id']
    conn = get_db_connection()
    cart_items = []
    total = 0
    
    if conn:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT c.cart_id, c.quantity, f.food_id, f.name, f.price, f.image_url, f.image_type,
                   (c.quantity * f.price) as subtotal
            FROM cart c
            JOIN food_items f ON c.food_id = f.food_id
            WHERE c.user_id = %s
        """, (user_id,))
        cart_items = cursor.fetchall()
        total = sum(item['subtotal'] for item in cart_items)
        # convert Decimal totals to float for template arithmetic
        try:
            from decimal import Decimal
            if isinstance(total, Decimal):
                total = float(total)
        except Exception:
            pass
        cursor.close()
        conn.close()
    
    return render_template('cart.html', cart_items=cart_items, total=total)

@app.route('/update_cart', methods=['POST'])
@login_required
def update_cart():
    cart_id = request.form['cart_id']
    quantity = int(request.form['quantity'])
    
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        if quantity > 0:
            cursor.execute("UPDATE cart SET quantity = %s WHERE cart_id = %s", (quantity, cart_id))
        else:
            cursor.execute("DELETE FROM cart WHERE cart_id = %s", (cart_id,))
        conn.commit()
        cursor.close()
        conn.close()
    
    return redirect(url_for('cart'))

@app.route('/remove_from_cart/<int:cart_id>')
@login_required
def remove_from_cart(cart_id):
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM cart WHERE cart_id = %s", (cart_id,))
        conn.commit()
        cursor.close()
        conn.close()
    
    return redirect(url_for('cart'))

@app.route('/checkout', methods=['GET', 'POST'])
@login_required
def checkout():
    if request.method == 'GET':
        user_id = session['user_id']
        conn = get_db_connection()
        cart_items = []
        total = 0
        
        if conn:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("""
                SELECT c.*, f.name, f.price, f.image_url, f.image_type, (c.quantity * f.price) as subtotal
                FROM cart c
                JOIN food_items f ON c.food_id = f.food_id
                WHERE c.user_id = %s
            """, (user_id,))
            cart_items = cursor.fetchall()
            total = sum(item['subtotal'] for item in cart_items)
            # convert Decimal to float for template arithmetic
            try:
                from decimal import Decimal
                if isinstance(total, Decimal):
                    total = float(total)
            except Exception:
                pass
            cursor.close()
            conn.close()
        
        return render_template('checkout.html', cart_items=cart_items, total=total)
    
    # POST request - process checkout
    user_id = session['user_id']
    payment_method = request.form.get('payment_method', 'Cash')
    order_type = request.form.get('order_type', 'immediate')
    scheduled_date = request.form.get('scheduled_date')
    scheduled_time = request.form.get('scheduled_time')
    is_bulk = request.form.get('is_bulk_order') == 'yes'
    event_name = request.form.get('event_name', '')
    
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor(dictionary=True)
        
        # Get cart items
        cursor.execute("""
            SELECT c.food_id, c.quantity, f.price
            FROM cart c
            JOIN food_items f ON c.food_id = f.food_id
            WHERE c.user_id = %s
        """, (user_id,))
        cart_items = cursor.fetchall()
        
        if not cart_items:
            flash('Your cart is empty!', 'error')
            return redirect(url_for('cart'))
        

        # Calculate total (may be Decimal)
        total_amount = sum(item['quantity'] * item['price'] for item in cart_items)

        # Optional special instructions from user (e.g., 'less sugar', 'extra spicy')
        special_instructions = request.form.get('special_instructions', '')

        # convert to float before applying float discounts to avoid Decimal*float errors
        try:
            from decimal import Decimal
            if isinstance(total_amount, Decimal):
                total_amount = float(total_amount)
        except Exception:
            pass

        # Get user type for priority
        cursor.execute("SELECT user_type, subscription_status FROM users WHERE user_id = %s", (user_id,))
        user = cursor.fetchone()
        priority = 'high' if user and user.get('user_type') == 'staff' else 'normal'

        # Apply subscription discount if applicable
        if user and user.get('subscription_status') == 'active':
            total_amount = total_amount * 0.9  # 10% discount
        
        # Create order - try to include special_instructions if DB supports it
        try:
            cursor.execute("""
                INSERT INTO orders (user_id, total_amount, payment_method, payment_status, status, 
                                  order_type, scheduled_date, scheduled_time, priority, is_bulk_order, event_name, special_instructions)
                VALUES (%s, %s, %s, 'Paid', 'placed', %s, %s, %s, %s, %s, %s, %s)
            """, (user_id, total_amount, payment_method, order_type, scheduled_date, scheduled_time, priority, is_bulk, event_name, special_instructions))
        except Error:
            # Fallback if `special_instructions` column doesn't exist
            cursor.execute("""
                INSERT INTO orders (user_id, total_amount, payment_method, payment_status, status, 
                                  order_type, scheduled_date, scheduled_time, priority, is_bulk_order, event_name)
                VALUES (%s, %s, %s, 'Paid', 'placed', %s, %s, %s, %s, %s, %s)
            """, (user_id, total_amount, payment_method, order_type, scheduled_date, scheduled_time, priority, is_bulk, event_name))

        order_id = cursor.lastrowid
        
        # Add order items
        for item in cart_items:
            cursor.execute("""
                INSERT INTO order_items (order_id, food_id, quantity, price)
                VALUES (%s, %s, %s, %s)
            """, (order_id, item['food_id'], item['quantity'], item['price']))
        
        # Clear cart
        cursor.execute("DELETE FROM cart WHERE user_id = %s", (user_id,))
        
        conn.commit()
        
        # Create notification
        create_notification(user_id, order_id, 'Order Placed', 
                          f'Your order #{order_id} has been placed successfully!')
        
        cursor.close()
        conn.close()
        
        flash('Order placed successfully!', 'success')
        return redirect(url_for('order_receipt', order_id=order_id))
    
    flash('Checkout failed!', 'error')
    return redirect(url_for('cart'))

@app.route('/cancel_order/<int:order_id>', methods=['POST'])
@login_required
def cancel_order(order_id):
    user_id = session['user_id']
    conn = get_db_connection()
    
    if conn:
        cursor = conn.cursor()
        try:
            # Check current order and payment status
            cursor.execute("SELECT status, payment_status FROM orders WHERE order_id = %s AND user_id = %s", (order_id, user_id))
            order = cursor.fetchone()
            if not order:
                return jsonify({'success': False, 'message': 'Order not found.'})

            current_status = order[0]
            payment_status = order[1]

            # Only allow cancellation before the 'preparing' stage
            non_cancellable = ('preparing', 'ready', 'completed', 'cancelled')
            if current_status in non_cancellable:
                return jsonify({'success': False, 'message': 'Order cannot be cancelled at this stage.'})

            # Update order to cancelled
            cursor.execute("""
                UPDATE orders
                SET status = 'cancelled'
                WHERE order_id = %s AND user_id = %s
            """, (order_id, user_id))

            # If payment was already paid, mark as refunded (simple flag) and notify user
            refund_message = None
            if payment_status == 'paid':
                cursor.execute("""
                    UPDATE orders SET payment_status = 'refunded' WHERE order_id = %s
                """, (order_id,))
                refund_message = 'Refunded successfully!'

            if cursor.rowcount > 0:
                conn.commit()
                # Notifications
                create_notification(user_id, order_id, 'Order Cancelled', f'Your order #{order_id} has been cancelled.')
                if refund_message:
                    create_notification(user_id, order_id, 'Refund Processed', f'A refund for order #{order_id} has been processed.')

                response_msg = refund_message if refund_message else 'Order cancelled successfully.'
                return jsonify({'success': True, 'message': response_msg})
            else:
                return jsonify({'success': False, 'message': 'Order could not be cancelled.'})
        except Error as e:
            return jsonify({'success': False, 'message': str(e)})
        finally:
            cursor.close()
            conn.close()
    
    return jsonify({'success': False, 'message': 'Database connection failed'})

@app.route('/reorder/<int:order_id>', methods=['POST'])
@login_required
def reorder(order_id):
    user_id = session['user_id']
    conn = get_db_connection()
    
    if conn:
        cursor = conn.cursor(dictionary=True)
        try:
            # Get items from the old order
            cursor.execute("""
                SELECT food_id, quantity
                FROM order_items
                WHERE order_id = %s
            """, (order_id,))
            order_items = cursor.fetchall()
            
            if not order_items:
                return jsonify({'success': False, 'message': 'No items found in this order.'})

            # Add items to cart
            for item in order_items:
                cursor.execute("""
                    INSERT INTO cart (user_id, food_id, quantity) 
                    VALUES (%s, %s, %s)
                    ON DUPLICATE KEY UPDATE quantity = quantity + VALUES(quantity)
                """, (user_id, item['food_id'], item['quantity']))
            
            conn.commit()
            return jsonify({'success': True})
        except Error as e:
            return jsonify({'success': False, 'message': str(e)})
        finally:
            cursor.close()
            conn.close()
    
    return jsonify({'success': False, 'message': 'Database connection failed'})

@app.route('/order_receipt/<int:order_id>')
@login_required
def order_receipt(order_id):
    user_id = session['user_id']
    conn = get_db_connection()
    order = None
    items = []
    
    if conn:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT o.*, u.name as user_name, u.email, u.phone
            FROM orders o
            JOIN users u ON o.user_id = u.user_id
            WHERE o.order_id = %s AND o.user_id = %s
        """, (order_id, user_id))
        order = cursor.fetchone()
        
        if order:
            cursor.execute("""
                SELECT oi.*, f.name
                FROM order_items oi
                JOIN food_items f ON oi.food_id = f.food_id
                WHERE oi.order_id = %s
            """, (order_id,))
            items = cursor.fetchall()

            # Convert scheduled_time if it's returned as a timedelta
            if order.get('scheduled_time') is not None:
                st = order['scheduled_time']
                try:
                    if isinstance(st, timedelta):
                        order['scheduled_time'] = (datetime.min + st).time()
                except Exception:
                    pass
        
        cursor.close()
        conn.close()
    
    return render_template('receipt.html', order=order, items=items)

@app.route('/download_receipt/<int:order_id>')
@login_required
def download_receipt(order_id):
    # Simple text receipt
    user_id = session['user_id']
    conn = get_db_connection()
    
    if conn:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT o.*, u.name, u.email
            FROM orders o
            JOIN users u ON o.user_id = u.user_id
            WHERE o.order_id = %s AND o.user_id = %s
        """, (order_id, user_id))
        order = cursor.fetchone()
        
        cursor.execute("""
            SELECT oi.*, f.name
            FROM order_items oi
            JOIN food_items f ON oi.food_id = f.food_id
            WHERE oi.order_id = %s
        """, (order_id,))
        items = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # Generate receipt text
        receipt = f"""
        ====================================
        CAMPUS EATS - DIGITAL RECEIPT
        ====================================
        Order ID: {order['order_id']}
        Date: {order['order_date']}
        Customer: {order['name']}
        Email: {order['email']}
        ====================================
        ITEMS:
        """
        
        for item in items:
            receipt += f"\n{item['name']} x {item['quantity']} - ₹{item['price'] * item['quantity']}"
        
        receipt += f"""
        
        ====================================
        Total Amount: ₹{order['total_amount']}
        Payment Method: {order['payment_method']}
        Payment Status: {order['payment_status']}
        ====================================
        Thank you for your order!
        """
        
        # Create BytesIO object
        buffer = BytesIO()
        buffer.write(receipt.encode('utf-8'))
        buffer.seek(0)
        
        return send_file(buffer, as_attachment=True, download_name=f'receipt_{order_id}.txt', mimetype='text/plain')
    
    flash('Receipt not found!', 'error')
    return redirect(url_for('orders'))

@app.route('/orders')
@login_required
def orders():
    user_id = session['user_id']
    conn = get_db_connection()
    orders_list = []
    
    if conn:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT o.*, COUNT(oi.order_item_id) as item_count
            FROM orders o
            LEFT JOIN order_items oi ON o.order_id = oi.order_id
            WHERE o.user_id = %s
            GROUP BY o.order_id
            ORDER BY o.order_date DESC
        """, (user_id,))
        orders_list = cursor.fetchall()
        
        for order in orders_list:
            # Parse stored JSON stage times into dict
            if order.get('stage_times'):
                order['stage_times'] = json.loads(order['stage_times'])

            # MySQL TIME fields may be returned as datetime.timedelta by the
            # mysql-connector; convert them to a datetime.time so templates
            # can safely call .strftime().
            if order.get('scheduled_time') is not None:
                st = order['scheduled_time']
                try:
                    if isinstance(st, timedelta):
                        order['scheduled_time'] = (datetime.min + st).time()
                    # leave as-is for other types (datetime.time or str)
                except Exception:
                    # If conversion fails, leave the raw value; template
                    # should handle or show nothing.
                    pass
        
        cursor.close()
        conn.close()
    
    return render_template('orders.html', orders=orders_list)

@app.route('/order_details/<int:order_id>')
@login_required
def order_details(order_id):
    user_id = session['user_id']
    conn = get_db_connection()
    order = None
    items = []
    stage_times = {}
    
    if conn:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT * FROM orders WHERE order_id = %s AND user_id = %s
        """, (order_id, user_id))
        order = cursor.fetchone()
        
        if order:
            cursor.execute("""
                SELECT oi.*, f.name, f.image_url, f.image_type
                FROM order_items oi
                JOIN food_items f ON oi.food_id = f.food_id
                WHERE oi.order_id = %s
            """, (order_id,))
            items = cursor.fetchall()
            
            # Parse stage times
            if order.get('stage_times'):
                stage_times = json.loads(order['stage_times'])

            # Prepare date/time fields for JSON response: convert to ISO/strings
            try:
                if isinstance(order.get('order_date'), datetime):
                    order['order_date'] = order['order_date'].isoformat()
            except Exception:
                pass

            if order.get('scheduled_date') is not None:
                try:
                    if isinstance(order['scheduled_date'], date):
                        order['scheduled_date'] = order['scheduled_date'].isoformat()
                except Exception:
                    pass

            if order.get('scheduled_time') is not None:
                st = order['scheduled_time']
                try:
                    if isinstance(st, timedelta):
                        # format as HH:MM:SS
                        order['scheduled_time'] = (datetime.min + st).time().strftime('%H:%M:%S')
                    elif hasattr(st, 'strftime'):
                        order['scheduled_time'] = st.strftime('%H:%M:%S')
                except Exception:
                    pass
        
        cursor.close()
        conn.close()
    
    return jsonify({'order': order, 'items': items, 'stage_times': stage_times})

@app.route('/rate_food/<int:food_id>', methods=['POST'])
@login_required
def rate_food(food_id):
    user_id = session['user_id']
    rating = int(request.form['rating'])
    review = request.form.get('review', '')
    order_id = request.form.get('order_id')
    
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        try:
            # Insert rating
            cursor.execute("""
                INSERT INTO ratings (user_id, food_id, order_id, rating, review)
                VALUES (%s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE rating = VALUES(rating), review = VALUES(review)
            """, (user_id, food_id, order_id, rating, review))
            
            # Update average rating
            cursor.execute("""
                UPDATE food_items 
                SET avg_rating = (SELECT AVG(rating) FROM ratings WHERE food_id = %s),
                    total_ratings = (SELECT COUNT(*) FROM ratings WHERE food_id = %s)
                WHERE food_id = %s
            """, (food_id, food_id, food_id))
            
            conn.commit()
            # If this is an AJAX request (from the rating modal), return JSON so the frontend can stay on the page
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': True, 'message': 'Rating submitted successfully!'})

            flash('Rating submitted successfully!', 'success')
        except Error as e:
            # Handle AJAX error response
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': f'Rating failed: {str(e)}'}), 500

            flash(f'Rating failed: {str(e)}', 'error')
        finally:
            cursor.close()
            conn.close()
    
    # Non-AJAX fallback: redirect back to orders page
    return redirect(url_for('orders'))

@app.route('/notifications')
@login_required
def notifications():
    user_id = session['user_id']
    conn = get_db_connection()
    notifications_list = []
    
    if conn:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT * FROM notifications 
            WHERE user_id = %s 
            ORDER BY created_at DESC 
            LIMIT 50
        """, (user_id,))
        notifications_list = cursor.fetchall()
        cursor.close()
        conn.close()
    
    return render_template('notifications.html', notifications=notifications_list)

@app.route('/mark_notification_read/<int:notification_id>', methods=['POST'])
@login_required
def mark_notification_read(notification_id):
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE notifications SET is_read = TRUE WHERE notification_id = %s
        """, (notification_id,))
        conn.commit()
        cursor.close()
        conn.close()
    return jsonify({'success': True})


@app.route('/clear_notifications', methods=['POST'])
@login_required
def clear_notifications():
    """Delete all notifications for the current user (clear history)."""
    user_id = session.get('user_id')
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        cursor.execute("""
            DELETE FROM notifications WHERE user_id = %s
        """, (user_id,))
        conn.commit()
        cursor.close()
        conn.close()
    return jsonify({'success': True})

@app.route('/get_unread_notifications')
@login_required
def get_unread_notifications():
    user_id = session['user_id']
    conn = get_db_connection()
    count = 0
    notifications_list = []
    
    if conn:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT COUNT(*) as count FROM notifications 
            WHERE user_id = %s AND is_read = FALSE
        """, (user_id,))
        count = cursor.fetchone()['count']
        
        cursor.execute("""
            SELECT * FROM notifications 
            WHERE user_id = %s AND is_read = FALSE 
            ORDER BY created_at DESC 
            LIMIT 5
        """, (user_id,))
        notifications_list = cursor.fetchall()
        
        cursor.close()
        conn.close()
    
    return jsonify({'count': count, 'notifications': notifications_list})

@app.route('/subscription')
@login_required
def subscription():
    user_id = session['user_id']
    conn = get_db_connection()
    user_data = None
    active_subscription = None
    
    if conn:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT * FROM users WHERE user_id = %s
        """, (user_id,))
        user_data = cursor.fetchone()
        
        cursor.execute("""
            SELECT * FROM subscriptions 
            WHERE user_id = %s AND status = 'active' 
            ORDER BY created_at DESC LIMIT 1
        """, (user_id,))
        active_subscription = cursor.fetchone()
        
        cursor.close()
        conn.close()
        # keep session subscription_status in sync
        try:
            session['subscription_status'] = user_data.get('subscription_status', 'inactive')
        except Exception:
            session['subscription_status'] = session.get('subscription_status', 'inactive')

    return render_template('subscription.html', user=user_data, subscription=active_subscription)

@app.route('/subscribe', methods=['POST'])
@login_required
def subscribe():
    user_id = session['user_id']
    plan_type = 'monthly'
    amount = 299.00  # Monthly subscription price
    
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        start_date = date.today()
        end_date = start_date + timedelta(days=30)
        
        # Create subscription
        cursor.execute("""
            INSERT INTO subscriptions (user_id, plan_type, amount, start_date, end_date, status)
            VALUES (%s, %s, %s, %s, %s, 'active')
        """, (user_id, plan_type, amount, start_date, end_date))
        
        # Update user subscription status
        cursor.execute("""
            UPDATE users 
            SET subscription_status = 'active', subscription_start = %s, subscription_end = %s
            WHERE user_id = %s
        """, (start_date, end_date, user_id))
        
        conn.commit()
        
        # Create notification
        create_notification(user_id, None, 'Subscription Activated', 
                          'Your monthly subscription has been activated! Enjoy 10% discount on all orders.')
        
        cursor.close()
        conn.close()
        
        flash('Subscription activated successfully!', 'success')
    # reflect subscription in session immediately
    session['subscription_status'] = 'active'
    
    return redirect(url_for('subscription'))

@app.route('/helpdesk', methods=['GET', 'POST'])
@login_required
def helpdesk():
    user_id = session['user_id']
    
    if request.method == 'POST':
        subject = request.form['subject']
        issue_type = request.form['issue_type']
        description = request.form['description']
        order_id_raw = request.form.get('order_id')

        # Open DB connection early so we can validate the optional order_id
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor(dictionary=True)

            # Normalize and validate order_id: must be numeric and refer to an existing order belonging to this user
            order_id = None
            if order_id_raw and order_id_raw.strip() != '':
                try:
                    candidate = int(order_id_raw)
                    cursor.execute("SELECT order_id FROM orders WHERE order_id = %s AND user_id = %s", (candidate, user_id))
                    found = cursor.fetchone()
                    if found:
                        order_id = candidate
                    else:
                        # invalid or not owned by user — ignore the provided order reference
                        flash('Provided order number was not found or does not belong to your account; ticket will be created without an order reference.', 'info')
                except ValueError:
                    flash('Invalid order number format; ticket will be created without an order reference.', 'info')

            # Insert ticket (order_id will be NULL if invalid/missing)
            # Use a non-dictionary cursor for INSERT to keep behavior consistent
            cursor.close()
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO helpdesk_tickets (user_id, order_id, subject, issue_type, description)
                VALUES (%s, %s, %s, %s, %s)
            """, (user_id, order_id, subject, issue_type, description))
            ticket_id = cursor.lastrowid
            conn.commit()

            # create a notification for the user so they know the ticket was created
            try:
                create_notification(user_id, None, 'Support Ticket Created', f'Your support ticket #{ticket_id} has been created')
            except Exception:
                # non-fatal: notification failure shouldn't block the user
                pass

            cursor.close()
            conn.close()

            flash('Support ticket created successfully!', 'success')
            return redirect(url_for('helpdesk'))
    
    # GET request - show tickets
    conn = get_db_connection()
    tickets = []
    
    if conn:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT * FROM helpdesk_tickets 
            WHERE user_id = %s 
            ORDER BY created_at DESC
        """, (user_id,))
        tickets = cursor.fetchall()
        cursor.close()
        conn.close()
    
    return render_template('helpdesk.html', tickets=tickets)

@app.route('/toggle_theme', methods=['POST'])
@login_required
def toggle_theme():
    user_id = session['user_id']
    theme = request.form['theme']
    
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE user_preferences SET theme = %s WHERE user_id = %s
        """, (theme, user_id))
        conn.commit()
        cursor.close()
        conn.close()
        session['theme'] = theme
    
    return jsonify({'success': True})

# ==================== ADMIN ROUTES ====================

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT * FROM admins WHERE username = %s AND password = %s", (username, password))
            admin = cursor.fetchone()
            cursor.close()
            conn.close()
            
            if admin:
                session['admin_id'] = admin['admin_id']
                session['admin_name'] = admin['username']
                return redirect(url_for('admin_dashboard'))
            else:
                flash('Invalid credentials!', 'error')
    
    return render_template('admin_login.html')

@app.route('/admin/logout')
def admin_logout():
    session.clear()
    return redirect(url_for('admin_login'))

@app.route('/admin/dashboard')
@admin_required
def admin_dashboard():
    conn = get_db_connection()
    stats = {}
    
    if conn:
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("SELECT COUNT(*) as count FROM users")
        stats['users'] = cursor.fetchone()['count']
        
        cursor.execute("SELECT COUNT(*) as count FROM food_items WHERE is_available = TRUE")
        stats['foods'] = cursor.fetchone()['count']
        
        cursor.execute("SELECT COUNT(*) as count FROM orders WHERE status IN ('placed', 'pending')")
        stats['pending_orders'] = cursor.fetchone()['count']
        
        cursor.execute("SELECT SUM(total_amount) as total FROM orders WHERE status = 'completed'")
        result = cursor.fetchone()
        stats['revenue'] = result['total'] if result['total'] else 0
        
        cursor.close()
        conn.close()
    
    return render_template('admin_dashboard.html', stats=stats)

@app.route('/admin/foods')
@admin_required
def admin_foods():
    conn = get_db_connection()
    foods = []
    categories = []
    
    if conn:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT f.*, c.category_name 
            FROM food_items f 
            LEFT JOIN categories c ON f.category_id = c.category_id 
            ORDER BY f.food_id DESC
        """)
        foods = cursor.fetchall()
        
        cursor.execute("SELECT * FROM categories")
        categories = cursor.fetchall()
        
        cursor.close()
        conn.close()
    
    return render_template('admin_foods.html', foods=foods, categories=categories)

@app.route('/admin/add_food', methods=['POST'])
@admin_required
def add_food():
    name = request.form['name']
    description = request.form['description']
    price = request.form['price']
    category_id = request.form['category_id']
    image_url = request.form.get('image_url', '/static/images/placeholder.jpg')
    image_type = request.form.get('image_type', 'local')
    
    # Handle file upload if local image
    if 'image_file' in request.files and request.files['image_file'].filename:
        file = request.files['image_file']
        filename = f"{datetime.now().timestamp()}_{file.filename}"
        file.save(os.path.join('static/images', filename))
        image_url = f'/static/images/{filename}'
        image_type = 'local'
    
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO food_items (name, description, price, category_id, image_url, image_type)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (name, description, price, category_id, image_url, image_type))
        conn.commit()
        cursor.close()
        conn.close()
        flash('Food item added successfully!', 'success')
    
    return redirect(url_for('admin_foods'))

@app.route('/admin/update_food/<int:food_id>', methods=['POST'])
@admin_required
def update_food(food_id):
    name = request.form['name']
    description = request.form['description']
    price = request.form['price']
    category_id = request.form['category_id']
    is_available = request.form.get('is_available', '0')
    
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE food_items 
            SET name = %s, description = %s, price = %s, category_id = %s, is_available = %s
            WHERE food_id = %s
        """, (name, description, price, category_id, is_available, food_id))
        conn.commit()
        cursor.close()
        conn.close()
        flash('Food item updated successfully!', 'success')
    
    return redirect(url_for('admin_foods'))

@app.route('/admin/delete_food/<int:food_id>')
@admin_required
def delete_food(food_id):
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM food_items WHERE food_id = %s", (food_id,))
        conn.commit()
        cursor.close()
        conn.close()
        flash('Food item deleted successfully!', 'success')
    
    return redirect(url_for('admin_foods'))

@app.route('/admin/orders')
@admin_required
def admin_orders():
    conn = get_db_connection()
    orders = []
    
    if conn:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT o.*, u.name as user_name, u.phone, u.user_type, COUNT(oi.order_item_id) as item_count
            FROM orders o
            JOIN users u ON o.user_id = u.user_id
            LEFT JOIN order_items oi ON o.order_id = oi.order_id
            GROUP BY o.order_id
            ORDER BY o.priority DESC, o.order_date DESC
        """)
        orders = cursor.fetchall()
        cursor.close()
        conn.close()
    
    return render_template('admin_orders.html', orders=orders)

@app.route('/admin/accept_order/<int:order_id>', methods=['POST'])
@admin_required
def accept_order(order_id):
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor(dictionary=True)
        
        # Update order status and set acceptance time
        cursor.execute("""
            UPDATE orders 
            SET status = 'pending', admin_accepted_at = NOW()
            WHERE order_id = %s
        """, (order_id,))

        # Generate stage times (5 min each)
        now = datetime.now()
        stage_times = {
            'placed': now.isoformat(),
            'pending': (now + timedelta(minutes=15)).isoformat(),
            'preparing': (now + timedelta(minutes=15)).isoformat(),
            'ready': (now + timedelta(minutes=15)).isoformat(),
            'completed': (now + timedelta(minutes=15)).isoformat()
        }
        
        cursor.execute("""
            UPDATE orders SET stage_times = %s WHERE order_id = %s
        """, (json.dumps(stage_times), order_id))
        
        # Get user_id for notification
        cursor.execute("SELECT user_id FROM orders WHERE order_id = %s", (order_id,))
        order = cursor.fetchone()
        
        conn.commit()
        cursor.close()
        conn.close()
        
        # Create notification
        create_notification(order['user_id'], order_id, 'Order Accepted', 
                          f'Your order #{order_id} has been accepted and is being prepared!')
        
        flash('Order accepted successfully!', 'success')
    
    return redirect(url_for('admin_orders'))

@app.route('/admin/update_order_status/<int:order_id>', methods=['POST'])
@admin_required
def update_order_status(order_id):
    status = request.form['status']
    
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor(dictionary=True)
        # Enforce status hierarchy (no reversing)
        cursor.execute("SELECT status FROM orders WHERE order_id = %s", (order_id,))
        current = cursor.fetchone()
        current_status = current['status'] if current and current.get('status') else None

        hierarchy = ['placed', 'pending', 'preparing', 'ready', 'completed', 'cancelled']
        try:
            current_index = hierarchy.index(current_status) if current_status in hierarchy else -1
        except ValueError:
            current_index = -1
        try:
            new_index = hierarchy.index(status) if status in hierarchy else -1
        except ValueError:
            new_index = -1

        if new_index < current_index:
            flash('Invalid status change: cannot reverse order status hierarchy.', 'error')
            cursor.close()
            conn.close()
            return redirect(url_for('admin_orders'))

        # Update status (allowed)
        cursor.execute("UPDATE orders SET status = %s WHERE order_id = %s", (status, order_id))
        
        # If completed, set completed_at
        if status == 'completed':
            cursor.execute("UPDATE orders SET completed_at = NOW() WHERE order_id = %s", (order_id,))
        
        # Get user_id for notification
        cursor.execute("SELECT user_id FROM orders WHERE order_id = %s", (order_id,))
        order = cursor.fetchone()
        
        conn.commit()
        cursor.close()
        conn.close()
        
        # Create notification
        status_messages = {
            'pending': 'Your order is pending',
            'preparing': 'Your order is being prepared',
            'ready': 'Your order is ready for pickup!',
            'completed': 'Your order has been completed',
            'cancelled': 'Your order has been cancelled'
        }
        
        create_notification(order['user_id'], order_id, 'Order Status Update', 
                          f'Order #{order_id}: {status_messages.get(status, "Status updated")}')
        
        flash('Order status updated!', 'success')
    
    return redirect(url_for('admin_orders'))

@app.route('/admin/users')
@admin_required
def admin_users():
    conn = get_db_connection()
    users = []
    
    if conn:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM users ORDER BY created_at DESC")
        users = cursor.fetchall()
        cursor.close()
        conn.close()
    
    return render_template('admin_users.html', users=users)

@app.route('/admin/delete_user/<int:user_id>')
@admin_required
def delete_user(user_id):
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM users WHERE user_id = %s", (user_id,))
        conn.commit()
        cursor.close()
        conn.close()
        flash('User deleted successfully!', 'success')
    
    return redirect(url_for('admin_users'))

@app.route('/admin/helpdesk')
@admin_required
def admin_helpdesk():
    conn = get_db_connection()
    tickets = []
    
    if conn:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT h.*, u.name as user_name, u.email
            FROM helpdesk_tickets h
            JOIN users u ON h.user_id = u.user_id
            ORDER BY h.created_at DESC
        """)
        tickets = cursor.fetchall()
        cursor.close()
        conn.close()
    
    return render_template('admin_helpdesk.html', tickets=tickets)

@app.route('/admin/respond_ticket/<int:ticket_id>', methods=['POST'])
@admin_required
def respond_ticket(ticket_id):
    response = request.form['response']
    status = request.form['status']
    
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            UPDATE helpdesk_tickets 
            SET admin_response = %s, status = %s, updated_at = NOW()
            WHERE ticket_id = %s
        """, (response, status, ticket_id))
        
        # Get user_id for notification
        cursor.execute("SELECT user_id FROM helpdesk_tickets WHERE ticket_id = %s", (ticket_id,))
        ticket = cursor.fetchone()
        
        conn.commit()
        cursor.close()
        conn.close()
        
        # Create notification
        create_notification(ticket['user_id'], None, 'Support Response', 
                          f'Your support ticket #{ticket_id} has been updated')
        
        flash('Response sent successfully!', 'success')
    
    return redirect(url_for('admin_helpdesk'))

@app.route('/admin/reports')
@admin_required
def admin_reports():
    return render_template('admin_reports.html')

@app.route('/admin/generate_report', methods=['POST'])
@admin_required
def generate_report():
    report_type = request.form['report_type']
    format_type = request.form['format']
    
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor(dictionary=True)
        
        # Calculate date range
        today = date.today()
        if report_type == 'daily':
            start_date = today
        elif report_type == 'weekly':
            start_date = today - timedelta(days=7)
        elif report_type == 'monthly':
            start_date = today - timedelta(days=30)
        else:  # quarterly
            start_date = today - timedelta(days=90)
        
        # Get report data
        cursor.execute("""
            SELECT 
                COUNT(*) as total_orders,
                SUM(total_amount) as total_revenue,
                AVG(total_amount) as avg_order_value,
                COUNT(DISTINCT user_id) as unique_customers
            FROM orders
            WHERE order_date >= %s AND status = 'completed'
        """, (start_date,))
        stats = cursor.fetchone()
        
        cursor.execute("""
            SELECT DATE(order_date) as date, COUNT(*) as orders, SUM(total_amount) as revenue
            FROM orders
            WHERE order_date >= %s AND status = 'completed'
            GROUP BY DATE(order_date)
            ORDER BY date
        """, (start_date,))
        daily_stats = cursor.fetchall()
        
        cursor.execute("""
            SELECT f.name, COUNT(oi.order_item_id) as total_sold, SUM(oi.quantity * oi.price) as revenue
            FROM order_items oi
            JOIN food_items f ON oi.food_id = f.food_id
            JOIN orders o ON oi.order_id = o.order_id
            WHERE o.order_date >= %s AND o.status = 'completed'
            GROUP BY f.food_id
            ORDER BY total_sold DESC
            LIMIT 10
        """, (start_date,))
        top_items = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # Generate report based on format
        if format_type == 'json':
            return jsonify({
                'report_type': report_type,
                'period': f'{start_date} to {today}',
                'summary': stats,
                'daily_breakdown': daily_stats,
                'top_items': top_items
            })

        # PDF generation
        if format_type == 'pdf':
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.pdfgen import canvas

                buffer = BytesIO()
                c = canvas.Canvas(buffer, pagesize=A4)
                width, height = A4
                margin = 50
                y = height - margin

                c.setFont('Helvetica-Bold', 16)
                c.drawString(margin, y, f'Campus Eats - {report_type.title()} Report')
                y -= 30

                c.setFont('Helvetica', 10)
                c.drawString(margin, y, f'Period: {start_date} to {today}')
                y -= 20

                # Summary
                c.setFont('Helvetica-Bold', 12)
                c.drawString(margin, y, 'Summary:')
                y -= 18
                c.setFont('Helvetica', 10)
                for k, v in (stats or {}).items():
                    c.drawString(margin + 10, y, f'{k}: {v}')
                    y -= 14
                    if y < margin:
                        c.showPage()
                        y = height - margin

                y -= 8
                # Daily breakdown
                c.setFont('Helvetica-Bold', 12)
                c.drawString(margin, y, 'Daily Breakdown:')
                y -= 18
                c.setFont('Helvetica', 10)
                if daily_stats:
                    for row in daily_stats:
                        line = f"{row.get('date')} - Orders: {row.get('orders')}, Revenue: {row.get('revenue')}"
                        c.drawString(margin + 10, y, line)
                        y -= 14
                        if y < margin:
                            c.showPage()
                            y = height - margin
                else:
                    c.drawString(margin + 10, y, 'No data')
                    y -= 14

                y -= 8
                # Top items
                c.setFont('Helvetica-Bold', 12)
                c.drawString(margin, y, 'Top Items:')
                y -= 18
                c.setFont('Helvetica', 10)
                if top_items:
                    for item in top_items:
                        line = f"{item.get('name')} - Sold: {item.get('total_sold')}, Revenue: {item.get('revenue')}"
                        c.drawString(margin + 10, y, line)
                        y -= 14
                        if y < margin:
                            c.showPage()
                            y = height - margin
                else:
                    c.drawString(margin + 10, y, 'No data')
                    y -= 14

                c.showPage()
                c.save()
                buffer.seek(0)
                return send_file(buffer, as_attachment=True, download_name=f'report_{report_type}_{start_date}.pdf', mimetype='application/pdf')
            except Exception as e:
                flash(f'PDF generation failed: {e}', 'error')
                return redirect(url_for('admin_reports'))

        # Excel generation
        if format_type in ('excel', 'xlsx'):
            try:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = 'Summary'

                ws.append(['Report Type', report_type])
                ws.append(['Period', f'{start_date} to {today}'])
                ws.append([])
                ws.append(['Summary'])
                if stats:
                    for k, v in stats.items():
                        ws.append([k, v])
                else:
                    ws.append(['No data'])

                ws2 = wb.create_sheet(title='Daily')
                ws2.append(['Date', 'Orders', 'Revenue'])
                for row in daily_stats or []:
                    ws2.append([row.get('date'), row.get('orders'), row.get('revenue')])

                ws3 = wb.create_sheet(title='Top Items')
                ws3.append(['Name', 'Total Sold', 'Revenue'])
                for it in top_items or []:
                    ws3.append([it.get('name'), it.get('total_sold'), it.get('revenue')])

                bio = BytesIO()
                wb.save(bio)
                bio.seek(0)
                return send_file(bio, as_attachment=True, download_name=f'report_{report_type}_{start_date}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            except Exception as e:
                flash(f'Excel generation failed: {e}', 'error')
                return redirect(url_for('admin_reports'))

        # Fallback: return JSON and a message
        flash('Report generation complete (JSON response).', 'success')
        return jsonify({
            'report_type': report_type,
            'summary': stats,
            'daily_breakdown': daily_stats,
            'top_items': top_items
        })
    
    flash('Report generation failed!', 'error')
    return redirect(url_for('admin_reports'))

@app.route('/admin/chart_data/<int:days>')
@admin_required
def chart_data(days):
    """Return JSON data for admin analytics charts."""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cursor = conn.cursor(dictionary=True)
        
        # Calculate date range
        end_date = datetime.now()
        start_date = end_date - timedelta(days=days)
        
        # Get daily orders and revenue
        cursor.execute("""
            SELECT DATE(order_date) as date,
                   COUNT(*) as order_count,
                   COALESCE(SUM(total_amount), 0) as revenue
            FROM orders 
            WHERE order_date >= %s AND order_date <= %s
            GROUP BY DATE(order_date)
            ORDER BY date
        """, (start_date.date(), end_date.date()))
        daily_stats = cursor.fetchall()
        
        # Create a complete date range with zeros for missing dates
        dates = []
        orders = []
        revenue = []
        
        # Create a lookup dictionary for easy access to stats
        stats_by_date = {str(stat['date']): stat for stat in daily_stats}
        
        # Fill in all dates in range
        current = start_date
        while current <= end_date:
            current_date = current.date()
            current_date_str = str(current_date)
            
            # Get stats for this date or use zeros
            stat = stats_by_date.get(current_date_str, {'order_count': 0, 'revenue': 0})
            
            dates.append(current_date.strftime('%Y-%m-%d'))
            orders.append(int(stat['order_count']))
            # Convert Decimal to float for JSON serialization
            try:
                revenue.append(float(stat['revenue']))
            except (TypeError, ValueError):
                revenue.append(0.0)
            
            current += timedelta(days=1)

        # Get popular items with error handling for NULL values
        cursor.execute("""
            SELECT f.name, 
                   COUNT(oi.order_item_id) as count
            FROM order_items oi
            JOIN food_items f ON oi.food_id = f.food_id
            JOIN orders o ON oi.order_id = o.order_id
            WHERE o.order_date >= %s AND o.order_date <= %s
            GROUP BY f.food_id, f.name
            ORDER BY count DESC
            LIMIT 10
        """, (start_date.date(), end_date.date()))
        popular_items = [
            {'name': str(row['name']), 'count': int(row['count'])}
            for row in cursor.fetchall()
        ]

        # Get order status distribution with error handling
        cursor.execute("""
            SELECT COALESCE(status, 'unknown') as status,
                   COUNT(*) as count
            FROM orders
            WHERE order_date >= %s AND order_date <= %s
            GROUP BY status
        """, (start_date.date(), end_date.date()))
        status_rows = cursor.fetchall()
        status_distribution = {
            str(row['status']): int(row['count'])
            for row in status_rows
        }

        cursor.close()
        conn.close()

        return jsonify({
            'dates': dates,
            'orders': orders,
            'revenue': revenue,
            'popular_items': popular_items,
            'status_distribution': status_distribution,
            'success': True
        })

    except Error as e:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    # Create images directory if it doesn't exist
    os.makedirs('static/images', exist_ok=True)
    app.run(debug=True, host='0.0.0.0', port=5000)
    